import openpyxl as xl
import datetime
import operator
import numpy as np
import sys
import collections
from os import listdir
from os.path import isfile, join

from scipy.constants import value
from sklearn.decomposition import PCA
from wheel.signatures.djbec import l


def has_duplicates(list_of_values):
    key_map = collections.defaultdict(int)
    for item in list_of_values:
        key_map[item] += 1

    for key in key_map.keys():
        if key_map[key] > 1:
            print("Key: " + str(key) + " Value: " + str(key_map[key]))
            return key
    return -1


start_time = datetime.datetime.now()

max_last_date = datetime.datetime(year=2017, month=12, day=23)
min_first_date = datetime.datetime(year=2017, month=9, day=8)

path = "data"

files = sorted(listdir(path))
N = len(files)
R = dict()
dates = []
i = 0

try:
    output = xl.load_workbook(path + "/" + "output.xlsx")
    output_ws = output.active
except FileNotFoundError:
    output = xl.Workbook()
    # output_ws = output.create_sheet(title="output")
    output_ws = output.active

output_z = output.create_sheet("Z")
output_ws["A1"] = "Ticker"

for file in files:
    wb = xl.load_workbook(path + "/" + file)
    draft_ws = wb["Draft"]
    output_ws["A" + str(i + 2)] = str(file[:-5])
    output_z.cell(row=1, column=i + 2).value = str(file[:-5])

    j = 0
    R[i] = []
    for row in draft_ws.rows:

        if min_first_date < row[0].value < max_last_date:
            if i == 0:
                dates.append(row[0].value)
            R[i].append(float(row[1].value))
            j += 1

    print(str(i) + ": len = " + str(len(R[i])) + " : " + file)
    i += 1

# Checking array's lengths
lenR = len(R[0])
for key in R.keys():
    if len(R[key]) != lenR:
        print("Error in array #" + str(key))
        print(R[key])
        print(len(R[key]))
        setRi = set(R[key])

# Filter working days
Len = len(R[0])
working_days = []
# checking for holiday
for i in range(0, Len):
    is_holiday = True
    for j in range(0, N):
        if R[j][i] != 0:
            is_holiday = False
    if not is_holiday:
        working_days.append(i)

n = len(dates)
m = 0
for day in range(0, n):
    if day in working_days:
        output_z.cell(row=m + 2, column=1).value = dates[day]
        m += 1

arr = []
R_buf = []
for i in range(0, Len):
    for j in range(0, N):
        R_buf.append([])
        for day in working_days:
            if i == day:
                R_buf[j].append(R[j][day])
print("Working days:", len(R_buf[0]))

R_ = np.array([R_buf[k] for k in range(0, N)])

# Z and sigma calculation
sigma = []
mu = []
z = []
t = 0

output_ws["B1"] = "Sigma"

for Ri in R_:
    mu_i = np.mean(Ri)
    mu.append(mu_i)

    sigma_i = np.var(Ri)
    output_ws["B" + str(t + 2)] = sigma_i
    sigma.append(sigma_i)

    buf = []
    l = 0
    for Rij in Ri:
        Zij = (Rij - mu_i) / sigma_i

        output_z.cell(row=l + 2, column=t + 2).value = Zij
        buf.append(Zij)
        l += 1
    z.append(buf)
    t += 1

# PCA
pca = PCA()
pca.fit(z)

ksi = pca.explained_variance_
L = pca.components_

lambda_ = ksi * ksi
omega_n = []
omega_n.append(lambda_[0])

for l in range(1, len(lambda_)):
    omega_n.append(omega_n[l-1] + lambda_[l])

h = []
output_ws["D1"] = "h"

output_ws["F1"] = "omega_n"
for o in range(0, len(omega_n)):
    output_ws["F" + str(o + 2)] = omega_n[o]

for o in range(0, len(omega_n)):
    h_n = omega_n[o]/omega_n[len(omega_n)-1]
    h.append(h_n)
    output_ws["D" + str(o +2)] = h_n

output_ws["C1"] = "Lambda"
for i in range(0, N):
    output_ws["C" + str(i + 2)] = lambda_[i]

print("z length: " + str(len(z)))
print("sigma length: " + str(len(sigma)))
print(sigma)
# Sigma_S calculation

sigma_S = 0
i = 0

for zi in z:
    j = 0

    for zj in z:
        mean = np.mean(np.array(zi) * np.array(zj))
        sigma_S += np.sqrt(sigma[j] * sigma[i]) * mean
        j += 1
    i += 1

# PCAS
PCAS = []
k = 1
i = 0

output_ws["E1"] = "PCAS"

for i in range(0, N):
    PCAS.append(0)
    for k in range(0, N):
        PCAS[i] += sigma[i] / sigma_S * L[k][i] * L[k][i] * lambda_[k]
    output_ws["E" + str(i + 2)] = PCAS[i]

print(PCAS)

output.save("output.xlsx")
end_time = datetime.datetime.now()
print()
print(end_time - start_time)
